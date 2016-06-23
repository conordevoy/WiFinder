
import openpyxl


# open the xlsx file to be operated on
wb = openpyxl.load_workbook('B0.02 B0.03 B0.04 Timetable.xlsx')


# make list of all rooms in the timetable sheet, excluding 'all'
# for future files, may be necessary to create a list of common exclusion sheets
# such as 'if i in excluded:'

rooms = []

for i in wb.get_sheet_names():
    if i == 'All':
        continue

    rooms.append(i)



# pull the first one to operate on

sheet = wb.get_sheet_by_name(rooms[0])



# print the value of a cell from the merged range list

sheet[sorted(sheet.merged_cell_ranges)[0][0:2]].value



# find first value in each cell merge range

mergecells = []

for i in range(len(sheet.merged_cell_ranges)):

    range_of_merge = sorted(sheet.merged_cell_ranges)[i]

    # only top left cell in merged cell contains a value, find this by splitting on ':' for 'A1:A2' = 'A1'
    top_left_cell_value = range_of_merge.split(":")[0]


    if sheet[top_left_cell_value].value == None:
        continue

    mergecells.append(top_left_cell_value)



# only keep mergecells from the 'module' rows

for row in sheet.iter_rows(row_offset=2):
    for cell in row:
        if cell.coordinate == 'N13': # local constraint, need to be checked for future datasets. Just to speed results.
            break
        if cell.coordinate in mergecells:

            cell = cell.coordinate
            cell_range = str(cell + ':' + cell[:1] ) + str( int(cell[1:]) + 1)

            sheet.unmerge_cells(cell_range)

            bottom_cell = cell_range.split(":")[1]
            sheet[bottom_cell].value = sheet[cell].value



wb.save('B002_finalmerge.xlsx')