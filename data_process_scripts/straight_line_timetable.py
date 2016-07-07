import openpyxl
import os

path_choice = '/home/mike/wifiproj/data' # local path, just to speed up script
os.chdir(path_choice)
print('The current working directory is:', os.getcwd())
path_choice = os.getcwd()
path_ok = input('To use another directory, please press any key followed by enter. Otherwise, just press enter. ')


if path_ok != '':
    path_choice = input('Please enter the directory to use: \n')
    invalid_path = True

    while invalid_path:
        try:
            os.chdir(path_choice)
            invalid_path = False
        except:
            path_choice = input('Invalid path. Please enter the directory to use: \n')


# open the xlsx file to be operated on
wb = openpyxl.load_workbook('B0.02 B0.03 B0.04 Timetable.xlsx')


# make list of all rooms in the timetable sheet, excluding 'all'
# for future files, may be necessary to create a list of common exclusion sheets
# such as 'if i in excluded:'

rooms = []

for sheetname in wb.get_sheet_names():
    if sheetname == 'All': # skip 'all' sheet
        continue

    rooms.append(sheetname) # append sheet to list


for room in rooms:

    # pull the sheet to operate on

    sheet = wb.get_sheet_by_name(room)



    # find first value in each cell merge range

    mergecells = [] # make list to store mergedcells

    for current_cell in range(len(sheet.merged_cell_ranges)): # move through each merge cell list

        range_of_merge = sorted(sheet.merged_cell_ranges)[current_cell]

        # only top left cell in merged cell contains a value, find this by splitting on ':' for 'A1:A2' = 'A1'
        top_left_cell_value = range_of_merge.split(":")[0]


        if sheet[top_left_cell_value].value == None: # skip any cells without data; some empty cells are merged
            continue

        mergecells.append(top_left_cell_value) # append any cell with data to the mergecells list



    # only keep mergecells from the 'module' rows

    for row in sheet.iter_rows(row_offset=2): # assumption; not all sheets may have the same row offset value
        for cell in row:

            if cell.coordinate in mergecells: # if the cell is a mergecell

                cell = cell.coordinate # take the coordinate of that mergecell, ie. 'X20'
                cell_range = str(cell + ':' + cell[:1] ) + str( int(cell[1:]) + 1) # make range: 'X20:X21'

                sheet.unmerge_cells(cell_range) # unmerge the cells

                bottom_cell = cell_range.split(":")[1] # make variable of cell 'X21'
                sheet[bottom_cell].value = sheet[cell].value # assign X20 = X21, bottom cell = top cell


    wb.save('.xlsx'.format(room)) # save each room as its own file

wb.save('CleanTimetable.xlsx') # save full timetable without double classes
